import os
import urllib.parse
import json
from datetime import datetime, date, timezone

import pandas as pd
import folium
from folium import FeatureGroup
from folium.plugins import Search
from openpyxl import load_workbook


# ------------------------
# Config
# ------------------------
EXCEL_FILE = "co_courses.xlsx"
OUTPUT_HTML = "index.html"
OUTPUT_JSON = "courses.json"
SCHEMA_VERSION = 2

TYPE_COLORS = {
    "Public": "#2ecc71",
    "Private": "#3498db",
    "Semi-Private": "#9b59b6",
    "Resort": "#f39c12",
}

DOT_RADIUS = 6
DOT_WEIGHT = 2


# ------------------------
# Helpers
# ------------------------
def normalize_type(x: str) -> str:
    if not isinstance(x, str) or not x.strip():
        return "Public"
    low = x.strip().lower()
    if low in ["semi private", "semi-private", "semi"]:
        return "Semi-Private"
    if low == "public":
        return "Public"
    if low == "private":
        return "Private"
    if low == "resort":
        return "Resort"
    return "Public"


def is_blank(v) -> bool:
    try:
        if pd.isna(v):
            return True
    except Exception:
        pass

    if v is None:
        return True
    if isinstance(v, str) and not v.strip():
        return True
    return False


def fmt_date(v) -> str:
    try:
        if pd.isna(v):
            return "—"
    except Exception:
        pass

    if v is None:
        return "—"

    if isinstance(v, pd.Timestamp):
        if pd.isna(v):
            return "—"
        v = v.to_pydatetime()

    if isinstance(v, (datetime, date)):
        try:
            return v.strftime("%-m/%-d/%Y")
        except Exception:
            return "—"

    s = str(v).strip()
    if not s:
        return "—"

    try:
        parsed = pd.to_datetime(s, errors="coerce")
        if pd.isna(parsed):
            return s
        return parsed.to_pydatetime().strftime("%-m/%-d/%Y")
    except Exception:
        return s


def to_iso_date_or_none(v):
    try:
        if pd.isna(v):
            return None
    except Exception:
        pass

    if v is None:
        return None

    if isinstance(v, pd.Timestamp):
        if pd.isna(v):
            return None
        v = v.to_pydatetime()

    if isinstance(v, (datetime, date)):
        try:
            return v.strftime("%Y-%m-%d")
        except Exception:
            return None

    s = str(v).strip()
    if not s:
        return None

    try:
        parsed = pd.to_datetime(s, errors="coerce")
        if pd.isna(parsed):
            return None
        return parsed.to_pydatetime().strftime("%Y-%m-%d")
    except Exception:
        return None


def to_int_or_none(v):
    if is_blank(v):
        return None
    try:
        return int(float(v))
    except Exception:
        s = str(v).strip()
        if not s:
            return None
        try:
            return int(s)
        except Exception:
            return None


def clean_text(s) -> str:
    if is_blank(s):
        return ""
    return str(s).strip()


def build_maps_links(address: str):
    q = urllib.parse.quote(address)
    google = f"https://www.google.com/maps/search/?api=1&query={q}"
    apple = f"https://maps.apple.com/?q={q}"
    return apple, google


def safe_js_str(s: str) -> str:
    return (
        str(s)
        .replace("\\", "\\\\")
        .replace('"', '\\"')
        .replace("\n", " ")
        .replace("\r", " ")
    )


# ------------------------
# Pull hyperlink targets from Excel for Reel column
# ------------------------
def extract_reel_links_xlsx(path: str, course_col_name="Course", reel_col_name="Reel"):
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if isinstance(v, str):
            headers[v.strip()] = c

    if course_col_name not in headers or reel_col_name not in headers:
        return {}

    course_col = headers[course_col_name]
    reel_col = headers[reel_col_name]

    out = {}
    for r in range(2, ws.max_row + 1):
        course = ws.cell(row=r, column=course_col).value
        if not isinstance(course, str) or not course.strip():
            continue
        course_name = course.strip()

        cell = ws.cell(row=r, column=reel_col)
        url = ""

        if cell.hyperlink and cell.hyperlink.target:
            url = str(cell.hyperlink.target).strip()

        if not url:
            v = cell.value
            if isinstance(v, str) and v.strip().lower().startswith("http"):
                url = v.strip()

        out[course_name] = url

    return out


# ------------------------
# Load + validate data
# ------------------------
df = pd.read_excel(EXCEL_FILE)
df = df.rename(columns=lambda c: c.strip())

required_cols = ["Course", "Address", "City", "Type", "Region", "Lat", "Long"]
missing = [c for c in required_cols if c not in df.columns]
if missing:
    raise ValueError(f"Missing columns: {missing}. Required: {required_cols}")

has_first_played = "1st Played" in df.columns
has_order = "Order" in df.columns
has_reel = "Reel" in df.columns

df["Type"] = df["Type"].apply(normalize_type)
df["Lat"] = pd.to_numeric(df["Lat"], errors="coerce")
df["Long"] = pd.to_numeric(df["Long"], errors="coerce")
df = df.dropna(subset=["Lat", "Long"]).copy()

reel_links = extract_reel_links_xlsx(EXCEL_FILE) if has_reel else {}

generated_dt = datetime.now(timezone.utc)
generated_at = generated_dt.isoformat().replace("+00:00", "Z")
generated_at_unix = int(generated_dt.timestamp())


# ------------------------
# Build map
# ------------------------
m = folium.Map(
    location=[39.0, -105.55],
    zoom_start=7,
    control_scale=True,
    tiles="OpenStreetMap",
)

MAP_JS_NAME = m.get_name()

type_groups = {}
type_group_js = {}
for t in TYPE_COLORS.keys():
    g = FeatureGroup(name=t, show=True, control=False)
    g.add_to(m)
    type_groups[t] = g
    type_group_js[t] = g.get_name()

markers_meta = []
courses_export = []

for _, r in df.iterrows():
    course = clean_text(r["Course"])
    city = clean_text(r["City"])
    ctype = normalize_type(r["Type"])
    region = clean_text(r["Region"])
    address = clean_text(r["Address"])

    first_played_display = fmt_date(r["1st Played"]) if has_first_played else "—"
    first_played_iso = to_iso_date_or_none(r["1st Played"]) if has_first_played else None
    played_by_cam = first_played_iso is not None

    order_num = to_int_or_none(r["Order"]) if has_order else None
    order_display = "—" if order_num is None else str(order_num)

    reel_url = ""
    if has_reel:
        reel_url = reel_links.get(course, "")
        if not reel_url:
            v = r.get("Reel", "")
            if isinstance(v, str) and v.strip().lower().startswith("http"):
                reel_url = v.strip()

    has_video = bool(reel_url)

    apple_maps, google_maps = build_maps_links(address if address else f"{course}, {city}, CO")
    color = TYPE_COLORS.get(ctype, TYPE_COLORS["Public"])

    if has_video:
        video_html = f"""
        <a href="{reel_url}" target="_blank" rel="noopener" style="display:block;text-decoration:none;">
          <div class="ec-video-btn">IG Reel</div>
        </a>
        """
    else:
        if played_by_cam:
            video_html = """
            <div class="ec-video-placeholder">No video yet</div>
            """
        else:
            video_html = """
            <div class="ec-video-placeholder ec-video-placeholder-muted">Not played yet</div>
            """

    popup_html = f"""
    <div class="ec-sheet-card" style="font-family:-apple-system,BlinkMacSystemFont,Segoe UI,Roboto,Arial;">
      <div class="ec-sheet-scroll">
        <div class="ec-sheet-topbar">
          <div class="ec-sheet-grabber"></div>
          <button class="ec-sheet-toggle" type="button" onclick="window.toggleBottomSheet && window.toggleBottomSheet()">
            ⌃
          </button>
        </div>

        <div class="ec-sheet-title">{course}</div>
        <div class="ec-sheet-subtitle">{city} · {region}</div>

        <div class="ec-info-row">
          <div class="ec-info-label">Type</div>
          <div class="ec-type-pill">
            <span class="ec-type-dot" style="background:{color};"></span>
            <span>{ctype}</span>
          </div>
        </div>

        <div class="ec-info-row ec-address-row">
          <div class="ec-info-label">Address</div>
          <div class="ec-info-value">{address}</div>
        </div>

        <div class="ec-btn-row">
          <a href="{apple_maps}" target="_blank" rel="noopener" class="ec-map-btn">Open in Maps</a>
          <a href="{google_maps}" target="_blank" rel="noopener" class="ec-map-btn">Google Maps</a>
        </div>

        <div class="ec-divider"></div>

        <div class="ec-section-title">Cam’s Every Course Journey</div>

        <div class="ec-journey-grid">
          <div class="ec-journey-card">
            <div class="ec-journey-label">Course #</div>
            <div class="ec-journey-value">{order_display}</div>
          </div>

          <div class="ec-journey-card">
            <div class="ec-journey-label">First Played</div>
            <div class="ec-journey-value">{first_played_display}</div>
            {video_html}
          </div>
        </div>
      </div>
    </div>
    """

    marker = folium.CircleMarker(
        location=[float(r["Lat"]), float(r["Long"])],
        radius=DOT_RADIUS,
        weight=DOT_WEIGHT,
        color=color,
        fill=True,
        fill_color=color,
        fill_opacity=0.9,
        popup=folium.Popup(popup_html, max_width=420),
    )
    marker.add_to(type_groups.get(ctype, type_groups["Public"]))

    markers_meta.append(
        {
            "js": marker.get_name(),
            "type": ctype,
            "played": played_by_cam,
            "video": has_video,
        }
    )

    courses_export.append(
        {
            "id": len(courses_export) + 1,
            "name": course,
            "city": city,
            "region": region,
            "type": ctype,
            "address": address,
            "lat": float(r["Lat"]),
            "lng": float(r["Long"]),
            "played": bool(played_by_cam),
            "order": order_num,
            "first_played": first_played_iso,
            "video_url": reel_url if reel_url else None,
            "has_video": bool(has_video),
            "apple_maps": apple_maps,
            "google_maps": google_maps,
        }
    )


# ------------------------
# Search layer (hidden GeoJSON)
# ------------------------
features = [
    {
        "type": "Feature",
        "properties": {"Course": clean_text(r["Course"])},
        "geometry": {"type": "Point", "coordinates": [float(r["Long"]), float(r["Lat"])]},
    }
    for _, r in df.iterrows()
]

search_layer = folium.GeoJson(
    {"type": "FeatureCollection", "features": features},
    name="__search_index__",
    show=False,
    control=False,
    marker=folium.CircleMarker(radius=0, opacity=0, fill_opacity=0),
    style_function=lambda x: {"opacity": 0, "fillOpacity": 0},
).add_to(m)

Search(
    layer=search_layer,
    search_label="Course",
    placeholder="Search a course name…",
    collapsed=False,
    position="topright",
    geom_type="Point",
    marker=False,
).add_to(m)


# ------------------------
# Filter UI
# ------------------------
filter_rows = "\n".join(
    [
        f"""
        <label class="ec-filter-option">
          <input type="checkbox" class="type-toggle" data-layer="{t}" checked>
          <span class="ec-filter-color" style="background:{TYPE_COLORS[t]};"></span>
          <span>{t}</span>
        </label>
        """
        for t in TYPE_COLORS.keys()
    ]
)

custom_css = """
<style>
  /* Search */
  .leaflet-control-search {
    z-index: 9997 !important;
    box-shadow: 0 6px 18px rgba(0,0,0,0.10) !important;
    border-radius: 12px !important;
    background: rgba(255,255,255,0.96) !important;
    border: 1px solid rgba(0,0,0,0.12) !important;
  }

  .leaflet-control-search .search-input {
    width: 220px !important;
    height: 36px !important;
    border-radius: 10px !important;
    font-size: 14px !important;
    padding-top: 0 !important;
    padding-bottom: 0 !important;
  }

  .leaflet-control-search .search-button {
    height: 36px !important;
  }

  /* Desktop popup tweaks */
  .leaflet-popup-content-wrapper {
    border-radius: 18px !important;
  }

  .leaflet-popup-content {
    margin: 12px 12px 14px !important;
    width: min(86vw, 420px) !important;
    max-width: min(86vw, 420px) !important;
  }

  /* Shared card styling */
  .ec-sheet-card {
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Arial, sans-serif;
  }

  .ec-sheet-scroll {
    height: 100%;
    overflow-y: auto;
    -webkit-overflow-scrolling: touch;
    padding-bottom: 12px;
  }

  .ec-sheet-topbar {
    position: relative;
    display: flex;
    justify-content: center;
    align-items: center;
    margin-bottom: 8px;
    min-height: 24px;
  }

  .ec-sheet-grabber {
    width: 42px;
    height: 5px;
    border-radius: 999px;
    background: rgba(0,0,0,0.16);
  }

  .ec-sheet-toggle {
    position: absolute;
    right: 0;
    top: -2px;
    border: 1px solid rgba(0,0,0,0.10);
    background: white;
    width: 30px;
    height: 30px;
    border-radius: 999px;
    font-size: 16px;
    line-height: 1;
    font-weight: 700;
    color: rgba(0,0,0,0.65);
    box-shadow: 0 2px 8px rgba(0,0,0,0.08);
    cursor: pointer;
  }

  .ec-sheet-title {
    font-weight: 900;
    font-size: 20px;
    line-height: 1.08;
    margin-bottom: 4px;
  }

  .ec-sheet-subtitle {
    font-size: 14px;
    opacity: 0.74;
    margin-bottom: 10px;
  }

  .ec-info-row {
    display: flex;
    align-items: flex-start;
    gap: 10px;
    margin-bottom: 8px;
  }

  .ec-address-row {
    margin-bottom: 10px;
  }

  .ec-info-label {
    width: 68px;
    opacity: 0.55;
    flex: 0 0 68px;
  }

  .ec-info-value {
    flex: 1;
    font-weight: 650;
    line-height: 1.35;
  }

  .ec-type-pill {
    display: inline-flex;
    align-items: center;
    gap: 8px;
    padding: 6px 10px;
    border-radius: 999px;
    background: rgba(0,0,0,0.04);
    font-weight: 800;
  }

  .ec-type-dot {
    width: 10px;
    height: 10px;
    border-radius: 3px;
    display: inline-block;
  }

  .ec-btn-row {
    display: flex;
    gap: 10px;
    margin: 10px 0 6px 0;
  }

  .ec-map-btn {
    flex: 1;
    padding: 10px 12px;
    border-radius: 14px;
    border: 1px solid rgba(0,0,0,0.16);
    text-align: center;
    font-weight: 900;
    color: #0b6aa2;
    background: white;
    text-decoration: none;
    box-sizing: border-box;
  }

  .ec-divider {
    height: 1px;
    background: rgba(0,0,0,0.12);
    margin: 12px 0;
  }

  .ec-section-title {
    font-weight: 900;
    font-size: 18px;
    margin-bottom: 8px;
  }

  .ec-journey-grid {
    display: grid;
    grid-template-columns: 1fr 1fr;
    gap: 10px;
  }

  .ec-journey-card {
    border: 1px solid rgba(0,0,0,0.12);
    border-radius: 14px;
    padding: 10px 12px;
  }

  .ec-journey-label {
    font-weight: 800;
    opacity: 0.6;
    margin-bottom: 6px;
  }

  .ec-journey-value {
    font-weight: 950;
    font-size: 20px;
    line-height: 1.1;
  }

  .ec-video-btn {
    width: 100%;
    margin-top: 10px;
    padding: 9px 10px;
    border-radius: 12px;
    border: 1px solid rgba(0,0,0,0.18);
    text-align: center;
    font-weight: 800;
    color: #0b6aa2;
    background: white;
    box-sizing: border-box;
  }

  .ec-video-placeholder {
    width: 100%;
    margin-top: 10px;
    padding: 9px 10px;
    border-radius: 12px;
    border: 1px solid rgba(0,0,0,0.12);
    text-align: center;
    font-weight: 800;
    color: rgba(0,0,0,0.45);
    background: rgba(0,0,0,0.03);
    box-sizing: border-box;
  }

  .ec-video-placeholder-muted {
    color: rgba(0,0,0,0.35);
    background: rgba(0,0,0,0.02);
  }

  /* Desktop filter box */
  .ec-filter-panel {
    position: fixed;
    bottom: 24px;
    right: 18px;
    z-index: 9998;
    background: rgba(255,255,255,0.94);
    border: 1px solid rgba(0,0,0,0.16);
    border-radius: 14px;
    padding: 12px 14px;
    box-shadow: 0 6px 20px rgba(0,0,0,0.10);
    font-family: -apple-system,BlinkMacSystemFont,Segoe UI,Roboto,Arial;
    font-size: 13px;
    width: 230px;
  }

  .ec-filter-title {
    font-weight: 900;
    font-size: 14px;
    margin-bottom: 8px;
  }

  .ec-filter-top {
    margin-bottom: 10px;
  }

  .ec-filter-top label {
    display: flex;
    align-items: center;
    gap: 10px;
    margin: 7px 0;
    cursor: pointer;
    font-weight: 800;
  }

  .ec-filter-divider {
    height: 1px;
    background: rgba(0,0,0,0.12);
    margin: 10px 0;
  }

  .ec-filter-option {
    display: flex;
    align-items: center;
    gap: 10px;
    margin: 7px 0;
    cursor: pointer;
  }

  .ec-filter-color {
    width: 14px;
    height: 14px;
    display: inline-block;
    border: 1px solid rgba(0,0,0,0.25);
  }

  .ec-filter-actions {
    display: flex;
    gap: 10px;
    margin-top: 10px;
  }

  .ec-filter-actions button {
    flex: 1;
    padding: 8px 10px;
    border-radius: 10px;
    border: 1px solid rgba(0,0,0,0.18);
    background: white;
    cursor: pointer;
  }

  /* Mobile-only controls */
  .ec-filter-toggle {
    display: none;
  }

  .ec-mobile-sheet {
    display: none;
  }

  .ec-mobile-backdrop {
    display: none;
  }

  /* Mobile overhaul */
  @media (max-width: 768px) {
    .leaflet-control-search {
      top: 10px !important;
      margin-top: 0 !important;
    }

    .leaflet-control-search .search-input {
      width: 190px !important;
      height: 34px !important;
      font-size: 13px !important;
    }

    .ec-filter-panel {
      display: none;
    }

    .ec-filter-toggle {
      display: inline-flex;
      position: fixed;
      right: 14px;
      bottom: 18px;
      z-index: 10002;
      align-items: center;
      justify-content: center;
      border-radius: 999px;
      border: 1px solid rgba(0,0,0,0.14);
      background: rgba(255,255,255,0.96);
      padding: 11px 16px;
      font-weight: 900;
      box-shadow: 0 6px 18px rgba(0,0,0,0.12);
      cursor: pointer;
      font-size: 14px;
    }

    .ec-mobile-backdrop.show {
      display: block;
      position: fixed;
      inset: 0;
      background: rgba(0,0,0,0.26);
      z-index: 10000;
    }

    .ec-mobile-sheet {
      display: block;
      position: fixed;
      left: 0;
      right: 0;
      bottom: -100%;
      z-index: 10001;
      background: white;
      border-radius: 18px 18px 0 0;
      box-shadow: 0 -8px 30px rgba(0,0,0,0.18);
      padding: 14px 16px 18px;
      transition: bottom 0.22s ease;
      max-height: 74vh;
      overflow-y: auto;
      font-family: -apple-system,BlinkMacSystemFont,Segoe UI,Roboto,Arial;
    }

    .ec-mobile-sheet.show {
      bottom: 0;
    }

    .ec-mobile-sheet-handle {
      width: 44px;
      height: 5px;
      border-radius: 999px;
      background: rgba(0,0,0,0.16);
      margin: 0 auto 12px;
    }

    .ec-mobile-sheet-title {
      font-weight: 900;
      font-size: 16px;
      margin-bottom: 10px;
    }

    .ec-mobile-sheet .ec-filter-top label,
    .ec-mobile-sheet .ec-filter-option {
      display: flex;
      align-items: center;
      gap: 10px;
      margin: 10px 0;
      cursor: pointer;
    }

    /* Hide default popup visuals on mobile */
    .leaflet-popup {
      opacity: 0 !important;
      pointer-events: none !important;
    }

    /* Bottom sheet */
    #ecBottomSheet {
      position: fixed;
      left: 0;
      right: 0;
      bottom: -100%;
      z-index: 10003;
      background: white;
      border-radius: 18px 18px 0 0;
      box-shadow: 0 -10px 34px rgba(0,0,0,0.22);
      transition: bottom 0.24s ease, height 0.22s ease;
      height: 170px;
      max-height: 70vh;
      overflow: hidden;
      padding: 10px 14px 14px;
      box-sizing: border-box;
    }

    #ecBottomSheet.show {
      bottom: 0;
    }

    #ecBottomSheet.expanded {
      height: 66vh;
    }

    #ecBottomSheet .ec-sheet-title {
      font-size: 18px;
    }

    #ecBottomSheet .ec-sheet-subtitle {
      font-size: 13px;
      margin-bottom: 10px;
    }

    #ecBottomSheet .ec-info-label {
      width: 58px;
      flex: 0 0 58px;
      font-size: 13px;
    }

    #ecBottomSheet .ec-info-value {
      font-size: 14px;
      line-height: 1.35;
    }

    #ecBottomSheet .ec-btn-row {
      gap: 8px;
    }

    #ecBottomSheet .ec-map-btn {
      padding: 9px 10px;
      font-size: 14px;
    }

    #ecBottomSheet .ec-section-title {
      font-size: 16px;
    }

    #ecBottomSheet .ec-journey-grid {
      grid-template-columns: 1fr 1fr;
      gap: 8px;
    }

    #ecBottomSheet .ec-journey-card {
      padding: 10px;
    }

    #ecBottomSheet .ec-journey-label {
      font-size: 12px;
    }

    #ecBottomSheet .ec-journey-value {
      font-size: 18px;
    }

    #ecBottomSheet .ec-video-btn,
    #ecBottomSheet .ec-video-placeholder {
      margin-top: 8px;
      padding: 8px 9px;
      font-size: 14px;
    }

    /* Collapsed state hides lower content */
    #ecBottomSheet:not(.expanded) .ec-divider,
    #ecBottomSheet:not(.expanded) .ec-section-title,
    #ecBottomSheet:not(.expanded) .ec-journey-grid {
      display: none;
    }

    #ecBottomSheet:not(.expanded) .ec-address-row {
      margin-bottom: 8px;
    }

    #ecBottomSheet:not(.expanded) .ec-btn-row {
      margin-top: 8px;
    }
  }
</style>
"""

ui_html = f"""
{custom_css}

<div class="ec-filter-panel" id="ecDesktopFilters">
  <div class="ec-filter-title">Course Filters</div>

  <div class="ec-filter-top">
    <label>
      <input id="playedOnly" type="checkbox">
      <span>Played by Cam</span>
    </label>

    <label>
      <input id="videoOnly" type="checkbox">
      <span>Has video review</span>
    </label>
  </div>

  <div class="ec-filter-divider"></div>

  {filter_rows}

  <div class="ec-filter-actions">
    <button id="filterAll">All</button>
    <button id="filterNone">None</button>
  </div>
</div>

<button class="ec-filter-toggle" id="ecFilterToggle">Filters</button>

<div class="ec-mobile-backdrop" id="ecMobileBackdrop"></div>

<div class="ec-mobile-sheet" id="ecMobileFilters">
  <div class="ec-mobile-sheet-handle"></div>
  <div class="ec-mobile-sheet-title">Course Filters</div>

  <div class="ec-filter-top">
    <label>
      <input id="playedOnlyMobile" type="checkbox">
      <span>Played by Cam</span>
    </label>

    <label>
      <input id="videoOnlyMobile" type="checkbox">
      <span>Has video review</span>
    </label>
  </div>

  <div class="ec-filter-divider"></div>

  {filter_rows.replace('class="type-toggle"', 'class="type-toggle-mobile"')}

  <div class="ec-filter-actions">
    <button id="filterAllMobile">All</button>
    <button id="filterNoneMobile">None</button>
  </div>
</div>

<div id="ecBottomSheet"></div>
"""

type_layers_js = ",\n".join([f'"{t}": window["{type_group_js[t]}"]' for t in TYPE_COLORS.keys()])

markers_js_list = ",\n".join(
    [
        f'{{m: window["{mm["js"]}"], type:"{safe_js_str(mm["type"])}", played:{str(mm["played"]).lower()}, video:{str(mm["video"]).lower()}}}'
        for mm in markers_meta
    ]
)

ui_js = f"""
document.addEventListener("DOMContentLoaded", function() {{
  const mapObj = window.{MAP_JS_NAME};
  if (!mapObj) {{
    console.warn("Map object not found");
    return;
  }}

  const typeLayers = {{
    {type_layers_js}
  }};

  const markers = [
    {markers_js_list}
  ];

  const bottomSheet = document.getElementById("ecBottomSheet");
  const filterToggle = document.getElementById("ecFilterToggle");
  const mobileFilters = document.getElementById("ecMobileFilters");
  const mobileBackdrop = document.getElementById("ecMobileBackdrop");

  const playedDesktop = document.getElementById("playedOnly");
  const videoDesktop = document.getElementById("videoOnly");
  const playedMobile = document.getElementById("playedOnlyMobile");
  const videoMobile = document.getElementById("videoOnlyMobile");

  let selectedLatLng = null;

  function isMobile() {{
    return window.innerWidth <= 768;
  }}

  function syncDesktopToMobile() {{
    if (playedMobile) playedMobile.checked = !!playedDesktop?.checked;
    if (videoMobile) videoMobile.checked = !!videoDesktop?.checked;

    const desktopToggles = Array.from(document.querySelectorAll(".type-toggle"));
    const mobileToggles = Array.from(document.querySelectorAll(".type-toggle-mobile"));

    mobileToggles.forEach((cb, idx) => {{
      if (desktopToggles[idx]) cb.checked = desktopToggles[idx].checked;
    }});
  }}

  function syncMobileToDesktop() {{
    if (playedDesktop) playedDesktop.checked = !!playedMobile?.checked;
    if (videoDesktop) videoDesktop.checked = !!videoMobile?.checked;

    const desktopToggles = Array.from(document.querySelectorAll(".type-toggle"));
    const mobileToggles = Array.from(document.querySelectorAll(".type-toggle-mobile"));

    desktopToggles.forEach((cb, idx) => {{
      if (mobileToggles[idx]) cb.checked = mobileToggles[idx].checked;
    }});
  }}

  function getTypeState() {{
    const state = {{}};
    const toggles = isMobile()
      ? document.querySelectorAll(".type-toggle-mobile")
      : document.querySelectorAll(".type-toggle");

    toggles.forEach(cb => {{
      state[cb.dataset.layer] = cb.checked;
    }});
    return state;
  }}

  function getPlayedOnly() {{
    return isMobile() ? !!playedMobile?.checked : !!playedDesktop?.checked;
  }}

  function getVideoOnly() {{
    return isMobile() ? !!videoMobile?.checked : !!videoDesktop?.checked;
  }}

  function applyFilters() {{
    const typeState = getTypeState();
    const playedOnly = getPlayedOnly();
    const videoOnly = getVideoOnly();

    Object.keys(typeLayers).forEach(t => {{
      const g = typeLayers[t];
      if (!g) return;

      const wantType = !!typeState[t];
      if (wantType) {{
        if (!mapObj.hasLayer(g)) mapObj.addLayer(g);
      }} else {{
        if (mapObj.hasLayer(g)) mapObj.removeLayer(g);
      }}
    }});

    markers.forEach(obj => {{
      if (!obj.m) return;

      let ok = !!typeState[obj.type];
      if (playedOnly) ok = ok && obj.played;
      if (videoOnly) ok = ok && obj.video;

      if (ok) {{
        if (!mapObj.hasLayer(obj.m)) mapObj.addLayer(obj.m);
      }} else {{
        if (mapObj.hasLayer(obj.m)) mapObj.removeLayer(obj.m);
      }}
    }});
  }}

  function openMobileFilters() {{
    mobileFilters.classList.add("show");
    mobileBackdrop.classList.add("show");
  }}

  function closeMobileFilters() {{
    mobileFilters.classList.remove("show");
    if (!bottomSheet.classList.contains("show")) {{
      mobileBackdrop.classList.remove("show");
    }}
  }}

  function openBottomSheet(html, latlng) {{
    if (!bottomSheet) return;
    bottomSheet.innerHTML = html || "";
    bottomSheet.classList.remove("expanded");
    bottomSheet.classList.add("show");
    mobileBackdrop.classList.add("show");
    selectedLatLng = latlng || null;
    focusSelectedOnMap(selectedLatLng, false);
    updateChevron();
  }}

  function closeBottomSheet() {{
    if (!bottomSheet) return;
    bottomSheet.classList.remove("show");
    bottomSheet.classList.remove("expanded");
    selectedLatLng = null;

    if (!mobileFilters.classList.contains("show")) {{
      mobileBackdrop.classList.remove("show");
    }}

    setTimeout(() => {{
      if (!bottomSheet.classList.contains("show")) {{
        bottomSheet.innerHTML = "";
      }}
    }}, 220);
  }}

  function updateChevron() {{
    const btn = bottomSheet?.querySelector(".ec-sheet-toggle");
    if (!btn) return;
    btn.textContent = bottomSheet.classList.contains("expanded") ? "⌄" : "⌃";
  }}

  function focusSelectedOnMap(latlng, expanded = false) {{
    if (!mapObj || !latlng) return;

    const currentZoom = mapObj.getZoom();
    const targetZoom = currentZoom < 10 ? 10 : currentZoom;

    mapObj.flyTo(latlng, targetZoom, {{
      animate: true,
      duration: 0.35
    }});

    setTimeout(() => {{
      const sheetHeight = expanded ? window.innerHeight * 0.66 : 170;
      const point = mapObj.project(latlng, mapObj.getZoom());
      const shiftedPoint = point.subtract([0, sheetHeight / 2.2]);
      mapObj.panTo(mapObj.unproject(shiftedPoint, mapObj.getZoom()), {{
        animate: true,
        duration: 0.25
      }});
    }}, 380);
  }}

  window.toggleBottomSheet = function() {{
    if (!bottomSheet) return;
    const expanded = bottomSheet.classList.toggle("expanded");
    updateChevron();
    if (selectedLatLng) {{
      setTimeout(() => {{
        focusSelectedOnMap(selectedLatLng, expanded);
      }}, 120);
    }}
  }};

  filterToggle?.addEventListener("click", () => {{
    syncDesktopToMobile();
    openMobileFilters();
  }});

  mobileBackdrop?.addEventListener("click", () => {{
    closeMobileFilters();
    closeBottomSheet();
  }});

  document.getElementById("filterAll")?.addEventListener("click", () => {{
    document.querySelectorAll(".type-toggle").forEach(cb => cb.checked = true);
    applyFilters();
  }});

  document.getElementById("filterNone")?.addEventListener("click", () => {{
    document.querySelectorAll(".type-toggle").forEach(cb => cb.checked = false);
    applyFilters();
  }});

  document.getElementById("filterAllMobile")?.addEventListener("click", () => {{
    document.querySelectorAll(".type-toggle-mobile").forEach(cb => cb.checked = true);
    syncMobileToDesktop();
    applyFilters();
  }});

  document.getElementById("filterNoneMobile")?.addEventListener("click", () => {{
    document.querySelectorAll(".type-toggle-mobile").forEach(cb => cb.checked = false);
    syncMobileToDesktop();
    applyFilters();
  }});

  playedDesktop?.addEventListener("change", applyFilters);
  videoDesktop?.addEventListener("change", applyFilters);

  playedMobile?.addEventListener("change", () => {{
    syncMobileToDesktop();
    applyFilters();
  }});

  videoMobile?.addEventListener("change", () => {{
    syncMobileToDesktop();
    applyFilters();
  }});

  document.querySelectorAll(".type-toggle").forEach(cb => {{
    cb.addEventListener("change", applyFilters);
  }});

  document.querySelectorAll(".type-toggle-mobile").forEach(cb => {{
    cb.addEventListener("change", () => {{
      syncMobileToDesktop();
      applyFilters();
    }});
  }});

  mapObj.on("popupopen", function(e) {{
    if (!isMobile()) return;

    setTimeout(() => {{
      const popupEl = e.popup && e.popup.getElement ? e.popup.getElement() : null;
      if (!popupEl) return;

      const contentEl = popupEl.querySelector(".leaflet-popup-content");
      if (!contentEl) return;

      openBottomSheet(contentEl.innerHTML, e.popup.getLatLng());

      setTimeout(() => {{
        try {{
          mapObj.closePopup();
        }} catch (err) {{
          console.warn(err);
        }}
      }}, 10);
    }}, 50);
  }});

  mapObj.on("click", function() {{
    if (isMobile()) closeBottomSheet();
  }});

  window.addEventListener("resize", () => {{
    if (!isMobile()) {{
      closeMobileFilters();
      closeBottomSheet();
    }}
    applyFilters();
  }});

  syncDesktopToMobile();
  applyFilters();
}});
"""

m.get_root().html.add_child(folium.Element(ui_html))
m.get_root().script.add_child(folium.Element(ui_js))


# ------------------------
# Save outputs next to this script
# ------------------------
base_dir = os.path.dirname(os.path.abspath(__file__))
out_html_path = os.path.join(base_dir, OUTPUT_HTML)
out_json_path = os.path.join(base_dir, OUTPUT_JSON)

m.save(out_html_path)

payload = {
    "meta": {
        "schema_version": SCHEMA_VERSION,
        "generated_at": generated_at,
        "generated_at_unix": generated_at_unix,
        "count": len(courses_export),
        "source_file": EXCEL_FILE,
    },
    "courses": courses_export,
}

with open(out_json_path, "w", encoding="utf-8") as f:
    json.dump(payload, f, ensure_ascii=False, indent=2)

print("Map created!")
print(f"Exported {len(courses_export)} rows to {out_json_path}")
print(f"generated_at: {generated_at} ({generated_at_unix})")